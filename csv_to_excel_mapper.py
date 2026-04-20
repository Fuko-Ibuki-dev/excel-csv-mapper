import openpyxl
import csv
import os
from openpyxl.utils import column_index_from_string
from datetime import datetime

# --------- Step 1: Load or create Excel ----------
excel_file = "./excel_csv_mapper/WPT_JR(B)26B1 - Copy.xlsx"

if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
else:
    wb = openpyxl.Workbook()
    sheet = wb.active

# --------- Step 2: Load CSV ----------
csv_file = "./excel_csv_mapper/WPT_JR(B)26B1_csv.csv"
csv_data = []
with open(csv_file, newline="", encoding="utf-8") as f:
    reader = csv.reader(f)
    csv_data = list(reader)

# --------- Step 3: User mapping ----------
mappings = [
    ("A2", "I4"),
    ("D2", "H4"),
    ("E2", "B4"),
    ("H2", "C4"),  # <-- this is the date column
    ("I2", "L4"),
    ("G2", "M4"),
]

# --------- Step 4: Transfer data dynamically ----------
for csv_cell, excel_cell in mappings:
    # Parse CSV source cell
    csv_col_letter = ''.join(filter(str.isalpha, csv_cell))
    csv_row = int(''.join(filter(str.isdigit, csv_cell)))
    csv_col = column_index_from_string(csv_col_letter) - 1  # 0-indexed for Python list

    # Parse Excel target cell
    excel_col_letter = ''.join(filter(str.isalpha, excel_cell))
    excel_row = int(''.join(filter(str.isdigit, excel_cell)))
    excel_col = column_index_from_string(excel_col_letter)

    # Transfer CSV column under source cell to Excel vertically
    for i in range(csv_row - 1, len(csv_data)):
        if csv_col < len(csv_data[i]):
            value = csv_data[i][csv_col]

            # --- Convert date string to datetime if this is a date column ---
            if csv_cell.upper() == "H2":  # check source cell
                try:
                    # parse "24/3/2026" to datetime
                    value = datetime.strptime(value, "%d/%m/%Y")
                except Exception as e:
                    print(f"⚠️ Could not parse date in {csv_cell}: {value}, writing as text")

            sheet.cell(row=excel_row + (i - (csv_row - 1)), column=excel_col, value=value)

# --------- Step 5: Save Excel ----------
wb.save(excel_file)
print("✅ CSV data transferred to Excel successfully!")